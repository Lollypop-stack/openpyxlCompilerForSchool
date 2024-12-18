import logging
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas import DataFrame


class Grade:
    def __init__(
        self,
        subjects: list["Subject"],
        grade: str,
        quarter: int | str,
    ) -> None:
        """
        Объект класса

        Args:
            subjects (list[Subject]): Список уроков
            grade (str): Класс (Пример: 3А)
            quarter (int | str): Четверть
        """

        self.logger = logging.getLogger(self.__class__.__name__)
        self.subjects = subjects
        self.grade = grade
        self.quarter = int(quarter)

    def to_excel(self, path: str = os.getcwd()) -> str:
        """
        Создание Excel файла с данными о предмете.

        Args:
            path (str, optional): Путь к директории для файла. Defaults to os.getcwd().

        Returns:
            str: Путь к файлу excel
        """
        normPath = Path(path)
        if not normPath.is_dir():
            raise ValueError("Это не директория")

        filePath = f"{normPath}\{self.grade}-{self.quarter}.xlsx"
        with pd.ExcelWriter(filePath, engine="xlsxwriter") as writer:
            for subject in self.subjects:
                subjectName = subject.name
                table = subject.table

                table.to_excel(writer, sheet_name=subjectName)
        self.logger.info(f"Данные сохранены в файл: {filePath}")
        return filePath

    def print_subjects(self) -> None:
        """
        Выввод уроков ввиде таблицы pandas
        """
        for subject in self.subjects:
            print(subject, end="\n")

    def __str__(self) -> str:
        return f"{self.grade}-{self.quarter}"


class Subject:
    def __init__(self, name: str, table: DataFrame) -> None:
        """
        Обьект урока

        Args:
            name (str): Название предмета
            table (DataFrame): Таблица pandas
        """
        self.name = name
        self.table = table

    def __str__(self) -> None:
        return f"{self.name}\n{self.table}"

    def get_students(self) -> list["Student"]:
        """
        Список студентов
        """
        students = [Student(i) for i in self.table.iloc[2:, 1].tolist()]
        return students

    def studentGrades(self, studentName: str) -> DataFrame:
        """
        Оценки ученика по предмету
        """
        student = self.table[self.table[1] == studentName]
        return student


class Student:
    def __init__(self, name: str) -> None:
        self.name = name


class GradeProccesor:
    def __init__(self, path: str) -> None:
        """
        Обработка класса
        Работает через with

        Args:
            path (str): путь к файлу excel
        """
        self.logger = logging.getLogger(self.__class__.__name__)
        self.path = Path(path).resolve()
        self.all_averages = {}
        self._active = False

        self.wb = None
        self.subjects = None
        self.result_sheet = None
        self.category_column = None
        self.averages_column = None

    def __enter__(self) -> "GradeProccesor":
        self._active = True

        if not self.path.exists():
            raise FileNotFoundError(
                f"Такого файла не существует: {self.path.as_posix()}"
            )
        else:
            try:
                with open(self.path, "a"):
                    pass
            except IOError:
                raise PermissionError("Файл открыт в другой программе")

        self.logger.info("Обработка файла excel")

        self.wb = load_workbook(self.path)

        self._create_result()

        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._active = False
        if exc_type is not None:
            info = (exc_type, exc_val, exc_tb)
            self.logger.error(f"Ошибка: ", exc_info=info)
        if self.wb:
            self.logger.info("Обработка окончена, запуск файла")

            # Сохраняем новый файл
            self.wb.save(self.path)
            os.startfile(self.path, "open")
        return False

    def _create_result(self) -> None:
        self.subjects = self.wb.sheetnames

        # Создаем лист для результатов
        if "Result" in self.subjects:
            self.wb.remove(self.wb["Result"])
            self.subjects = self.subjects[1:]
        self.result_sheet = self.wb.create_sheet("Result", 0)

        # Записываем название класса в объединенную ячейку A1
        self.result_sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(self.subjects) + 3
        )
        self.result_sheet["A1"] = f"{self.path.stem} результаты"

        # Устанавливаем заголовки "Ученики"
        self.result_sheet["B2"] = "Ученики"

        # Пройдемся по каждому листу, чтобы получить оценки
        for col, subject_name in enumerate(self.subjects, start=3):
            self.result_sheet.cell(row=2, column=col, value=subject_name)

            subject_sheet = self.wb[subject_name]

            # Поиск колонки с текстом "СР"
            sr_col = subject_sheet.max_column - 1

            # Копируем средние баллы для каждого ученика
            for i in range(4, subject_sheet.max_row + 1):
                student = subject_sheet.cell(row=i, column=2).value

                grade = subject_sheet.cell(row=i, column=sr_col).value

                all_grades = self.all_averages.get(student)
                if all_grades is None:
                    all_grades = [None for _ in self.subjects]

                all_grades[col - 3] = grade

                # Добавляем оценку в общий список
                self.all_averages.update({student: all_grades})

        # Заполняем средние оценки по всем предметам для каждого ученика
        for row, student in enumerate(self.all_averages.items(), start=3):
            studentName = student[0]
            studentGrades = student[1]

            studentRow = [row - 2, studentName]
            studentRow.extend(studentGrades)

            self.result_sheet.append(studentRow)

        self.category_column = len(self.subjects) + 4  # Колонка для записи категорий
        self.averages_column = len(self.subjects) + 3  # Колонка для записи общей оценки

        self.logger.info("Лист c результатами создан")

    def get_grade(self) -> Grade:
        """
        Создает объект класса из excel файла

        Returns:
            Grade: Объект класса
        """
        subjects = pd.read_excel(self.path, sheet_name=None, header=(0, 1))
        subjects = [
            Subject(i[0], i[1].iloc[:, 1:])
            for i in subjects.items()
            if i[0] != "Result"
        ]

        grade = self.path.stem.split("-")
        return Grade(subjects, grade[0], grade[1])

    def _isContext(func) -> None:
        def check(self):
            if not self._active:
                raise RuntimeError(
                    "Этот класс можно использовать только внутри контекстного менеджера(with)"
                )
            else:
                func(self)

        return check

    def start(self) -> None:
        """
        Запуск всех функций без with
        """
        with self as xl:
            xl.all()

    @_isContext
    def all(self) -> None:
        """
        Все функции
        """
        self.calculate_averages()
        self.assign_categories()
        self.create_category_pie()

    @_isContext
    def calculate_averages(self) -> None:
        """
        Создание колонны для среднего балла по всем предметам для учеников
        """

        # Заголовок для средней оценки по всем предметам
        self.result_sheet.cell(
            row=2, column=self.averages_column, value="Средний балл по всем предметам"
        )

        # Заполняем средние оценки по всем предметам для каждого ученика
        for row, student in enumerate(self.all_averages.values(), start=3):
            studentGrades = [i for i in student if type(i) == int or type(i) == float]

            if studentGrades:
                total = sum(studentGrades) / len(studentGrades)
            else:
                total = 0

            self.result_sheet.cell(row=row, column=self.averages_column, value=total)

        self.logger.info("Средний бал записан")

    @_isContext
    def assign_categories(self) -> None:
        """
        Добавляет категории (Отл., Уд., Тр., Дв., Нз.) рядом со средним баллом.
        """

        self.result_sheet.cell(row=2, column=self.category_column, value="Категория")

        for row in range(3, len(self.all_averages) + 3):
            avg_score = self.result_sheet.cell(row=row, column=self.category_column - 1).value
            if avg_score is None:
                raise ValueError("Коллонна с общими оценками отсутствует")

            if avg_score == 0:
                category = "Нз."  # Незачет, если средний балл отсутствует
            elif avg_score > 4.6:
                category = "Отл."  # Отличник
            elif avg_score > 3.6:
                category = "Уд."  # Ударник
            elif avg_score > 2.6:
                category = "Тр."  # Троечник
            else:
                category = "Дв."  # Двоечник

            self.result_sheet.cell(row=row, column=self.category_column, value=category)

        self.logger.info("Категории записанны")

    @_isContext
    def create_category_pie(self) -> None:
        """
        Создает круговую диаграмму по категориям студентов (Отл., Уд., Тр., Дв., Нз.)
        Мини-таблица с категориями будет отображаться справа от столбца с предметами.
        """
        # Подсчитываем количество студентов в каждой категории
        cat_counts = {"Отл.": 0, "Уд.": 0, "Тр.": 0, "Дв.": 0, "Нз.": 0}

        # Собираем данные о категориях
        for row in range(3, len(self.all_averages) + 3):
            category = self.result_sheet.cell(row=row, column=self.category_column).value
            if category in cat_counts:
                cat_counts[category] += 1

        # Записываем мини-таблицу с категориями
        table_start_col = self.category_column + 2
        table_start_row = 3  # Мини-таблица начинается с 3 строки
        self.result_sheet.cell(row=2, column=table_start_col, value="Категория")
        self.result_sheet.cell(row=2, column=table_start_col + 1, value="Кол-во")
        self.result_sheet.cell(row=2, column=table_start_col + 2, value="Процент")

        total_students = len(self.all_averages)
        for row, category in enumerate(cat_counts.items(), start=table_start_row):
            self.result_sheet.cell(row=row, column=table_start_col, value=category[0])
            self.result_sheet.cell(
                row=row, column=table_start_col + 1, value=category[1]
            )
            percent = category[1] / total_students * 100
            self.result_sheet.cell(
                row=row, column=table_start_col + 2, value=f"{percent:.2f}%"
            )

        # Создаем круговую диаграмму
        chart = PieChart()
        data = Reference(
            self.result_sheet,
            min_col=table_start_col + 1,
            min_row=table_start_row,
            max_row=table_start_row + len(cat_counts) - 1,
        )
        categories_reference = Reference(
            self.result_sheet,
            min_col=table_start_col,
            min_row=table_start_row,
            max_row=table_start_row + len(cat_counts) - 1,
        )

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories_reference)
        chart.title = "Распределение категорий"

        # Позиция диаграммы (под мини-таблицей)
        chart_anchor = f"{get_column_letter(table_start_col)}{table_start_row + len(cat_counts) + 1}"
        self.result_sheet.add_chart(chart, chart_anchor)

        self.logger.info("Пирог создан")
