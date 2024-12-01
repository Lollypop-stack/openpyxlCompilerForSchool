import os
import time
from io import StringIO
from threading import Thread
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pandas import DataFrame

from openpyxl import load_workbook
from pathlib import Path
import re

import tkinter as tk
from tkinter import messagebox


class RThread(Thread):
    """
    Расширение потока для выполнения задачи и хранения результата.
    """

    def __init__(self, target, args):
        Thread.__init__(self)
        self.target = target
        self.result = None

    def run(self) -> None:
        self.result = self.target()  # Выполнение целевой функции


class Grade:
    """
    Объект класса.

    Args:
        subjects (dict[str, DataFrame]): урок, таблица
        grade (str): класс
        quarter (int): четверть
    """

    def __init__(self, subjects: dict[str, DataFrame], grade: str, quarter: int):
        self.subjects = subjects
        self.grade = grade
        self.quarter = quarter

    def to_excel(self, path: str = "./data"):
        """
        Создание Excel файла с данными о предмете.
        """
        path = f"{path}/{self.grade}-{self.quarter}.xlsx"
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            for i in self.subjects.items():
                subject = i[0]
                table = i[1]

                table.to_excel(writer, sheet_name=subject)

        print(f'file "{path}" created')

    def print_subjects(self):
        """
        Выввод уроков ввиде таблицы pandas
        """
        for i in self.subjects.items():
            subject = i[0]
            table = i[1]

            print(subject, table, sep="\n", end="\n")


class KParser:
    """
    Класс для работы с системой Kundoluk.

    Основные функции:
    - Получение данных о классах (get_grade)
    - Получение данных по предметам (get_subject)
    - Обработка пользовательского ввода (magic)
    """

    def __init__(self) -> None:
        # Настраиваем заголовки для HTTP-запросов
        self.__headers = {
            "cookie": f"language=ru; session={os.getenv('SESSION')}; deferApp=true; string=670fbb44e3637",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "dnt": "1",
            "priority": "u=0, i",
            "^sec-ch-ua": r"^\^Not/A",
            "sec-ch-ua-mobile": "?0",
            "^sec-ch-ua-platform": r"^\^Windows^^^",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "sec-gpc": "1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6324.206 Safari/537.36",
        }
        self.url = "https://kundoluk.edu.kg/journal2"  # Базовый URL для запросов
        self.s = (
            requests.Session()
        )  # Создаем сессию для повторного использования соединений
        # Список классов и их идентификаторов
        self.grades = [
            ("2А", 90359),
            ("2Б", 90360),
            ("3А", 90340),
            ("3Б", 90341),
            ("4А", 90343),
            ("4Б", 90344),
            ("5А", 90345),
            ("5Б", 90346),
            ("6А", 90347),
            ("6Б", 90348),
            ("7А", 90349),
            ("7Б", 90350),
            ("8А", 90351),
            ("8Б", 90352),
            ("9А", 90353),
            ("9Б", 90354),
            ("10А", 90355),
            ("10Б", 90356),
            ("11А", 90357),
            ("11Б", 90358),
            ("2В", 90361),
        ]

    def get_grade(self, grade: tuple[str, int], quarter: int = 1) -> Grade:
        """
        Получение данных для конкретного класса и четверти.

        Args:
            grade (tuple[str, int]): Класс и его ID
            quarter (int): Номер четверти

        Returns:
            Grade: Объект Grade с данными класса
        """
        try:
            querystring = {"class": str(grade[1]), "quarter": str(quarter)}
            response = self.s.get(self.url, headers=self.__headers, params=querystring)
            response.raise_for_status()  # Проверка успешности запроса

            bs = BeautifulSoup(response.content, "lxml")
            subjects = bs.find("ul", class_="uk-subnav").find_all("a")
            threads = []
            for link in subjects:
                url = link["href"]
                name = link.text.strip()

                # Создаем поток для получения данных о предмете
                thread = RThread(target=lambda: self.get_subject(url), args=())
                thread.start()
                threads.append((thread, name))

                time.sleep(0.1)  # Небольшая пауза между запросами

            tables = []
            for thread in threads:
                while thread[0].result is None:  # Ожидание завершения потока
                    time.sleep(0.1)
                table = thread[0].result
                if type(table) != int:
                    tables.append((thread[1], table))
            # Сортировка по названию урока
            tables = dict(sorted(tables, key=lambda x: x[0]))

            return Grade(tables, grade[0], quarter)
        except requests.exceptions.RequestException as e:
            raise Exception(f"Ошибка при запросе данных для класса {grade[0]}: {e}")

    def get_subject(self, url: str) -> DataFrame:
        """
        Получение данных о предмете по ссылке.

        Args:
            url (str): Ссылка на предмет

        Returns:
            DataFrame: Таблица с оценками
        """
        response = self.s.request("GET", url, headers=self.__headers)
        # Проверка успешности запроса
        if not response.ok:
            raise Exception(
                "subjectError:",
                f"subject url({url}) response code {response.status_code}",
            )

        bs = BeautifulSoup(response.content, "lxml")

        try:
            htmlTable = bs.find(
                "table", class_="elementFixed-striped"
            )  # Таблица с данными
            # Удаляем лишние элементы из таблицы
            trash = htmlTable.find_all("span", class_="uk-margin-xsmall-right")
            [i.extract() for i in trash]
            table = pd.read_html(StringIO(str(htmlTable)))[
                0
            ]  # Конвертируем в DataFrame
        except Exception:
            return 10
        return table

    def magic(self) -> None:
        """
        Основная логика обработки пользовательского ввода из консоли.
        """
        try:
            while True:
                try:
                    inpGrade = input("Введите класс (Пример: 4Б): ").upper()
                    inpQuarter = input("Введите четверть: ").strip()

                    if not inpQuarter.isdigit() or not inpGrade:
                        print("Неверный ввод. Попробуйте снова.")
                        continue

                    inpQuarter = int(inpQuarter)
                    grades = dict(self.grades)
                    gradeId = grades.get(inpGrade)

                    if not gradeId or inpQuarter <= 0:
                        print("Класс или четверть указаны неверно. Попробуйте снова.")
                        continue

                    # Получение данных о классе
                    grade = self.get_grade((inpGrade, gradeId), inpQuarter)

                    # Убедитесь, что директория существует
                    output_dir = os.path.join(os.getcwd(), 'data')  # Папка 'data' рядом с программой
                    os.makedirs(output_dir, exist_ok=True)

                    # Создаем путь к файлам
                    input_file = os.path.join(output_dir, f'{inpGrade}-{inpQuarter}.xlsx')
                    grade.to_excel(input_file)

                    # Генерация output_file
                    output_file = os.path.join(output_dir, f'{inpGrade}-{inpQuarter}-result.xlsx')
                    print(f"Генерация файла: {output_file}")

                    # Проверка на существование output_file
                    if not output_file:
                        raise ValueError("Неверный путь для выходного файла.")

                    # Обрабатываем данные и сохраняем в новый файл
                    calculate_averages(input_file, output_file)

                    print(f'Результаты сохранены в файл: {output_file}')
                    break
                except Exception as e:
                    import traceback
                    print(f"Ошибка: {e}")
                    print("Traceback:")
                    print(traceback.format_exc())  # Вывод полного traceback
        except KeyboardInterrupt:
            print("\nПрограмма была завершена пользователем.")


def extract_digit(text):
    """Функция для извлечения чисел из текста, поддерживает как целые, так и дробные числа."""
    match = re.search(r'[\d,\.]+', text)
    return float(match.group().replace(',', '.')) if match else 0.0  # Обрабатываем как дробное число

def calculate_averages(input_file, output_file):
    wb = load_workbook(input_file)

    # Создаем лист для результатов
    result_sheet = wb.create_sheet('Result')


    # Записываем название класса в объединенную ячейку A1
    subjects = [subject for subject in wb.sheetnames if subject != 'Result']
    result_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subjects) + 3)

    # ////////////////////////////////////////////////
    input_path = Path(input_file)
    file_name = input_path.stem
    result_sheet['A1'] = f"{file_name} результаты"
    # ////////////////////////////////////////////////

    # Устанавливаем заголовки "Ученик" (объединенные B2 и C2)
    result_sheet.merge_cells('B2:C2')
    result_sheet['B2'] = 'Ученик'

    # Записываем названия предметов в строку 2
    for col, subject in enumerate(subjects, start=4):
        result_sheet.cell(row=2, column=col, value=subject)

    # Заголовок для средней оценки по всем предметам
    result_sheet.cell(row=2, column=len(subjects) + 4, value="Средний балл по всем предметам")

    # Получение списка учеников с первого листа
    subject_sheet = wb[subjects[0]]
    students = [
        f"{str(subject_sheet.cell(row=i, column=1).value)} {str(subject_sheet.cell(row=i, column=2).value)}"
        for i in range(4, subject_sheet.max_row + 1)
    ]

    # Массив для хранения всех оценок по всем предметам для каждого ученика
    all_averages = {student: [] for student in students}

    # Пройдемся по каждому листу, чтобы получить оценки
    for col, subject_name in enumerate(subjects, start=4):
        subject_sheet = wb[subject_name]

        # Поиск колонки с текстом "СР" (предположим, что она в 2-й строке)
        sr_col = None
        for col_idx in range(1, subject_sheet.max_column + 1):
            if subject_sheet.cell(row=2, column=col_idx).value == "СР":  # Ищем колонку "СР"
                sr_col = col_idx
                break

        # Если нашли колонку "СР", копируем оценки для всех учеников
        if sr_col:
            # Записываем название предмета в строку 3 в колонку результата
            result_sheet.cell(row=3, column=col, value=subject_name)

            # Копируем средние баллы для каждого ученика
            for i, student in enumerate(students, start=4):
                grade = subject_sheet.cell(row=i, column=sr_col).value
                grade = extract_digit(str(grade))  # Извлекаем число из строки, теперь поддерживаем дробные числа
                all_averages[student].append(grade)  # Добавляем оценку в общий список
                result_sheet.cell(row=i - 1, column=col, value=grade)  # Записываем оценку в итоговую таблицу (сдвиг на 1 вниз)

    # Заполняем средние оценки по всем предметам для каждого ученика
    for row, student in enumerate(students, start=3):
        result_sheet.cell(row=row, column=1, value=row - 2)  # Номер ученика
        result_sheet.cell(row=row, column=2, value=student)

        # Рассчитываем средний балл по всем предметам для каждого ученика (с округлением)
        total_average = sum(all_averages[student]) / len(subjects)
        result_sheet.cell(row=row, column=len(subjects) + 4, value=round(total_average, 2))  # Округление для среднего балла по всем предметам

    # Сохраняем новый файл
    wb.save(output_file)

def start_ui():
    # Код интерфейса
    root = tk.Tk()
    root.title("Kundoluk Parser")

    tk.Label(root, text="Введите класс (например, 4Б):").grid(row=0, column=0)
    tk.Label(root, text="Введите четверть (число):").grid(row=1, column=0)

    grade_entry = tk.Entry(root)
    quarter_entry = tk.Entry(root)

    grade_entry.grid(row=0, column=1)
    quarter_entry.grid(row=1, column=1)

    def submit():
        grade = grade_entry.get().strip().upper()
        quarter = quarter_entry.get().strip()
        if not grade or not quarter.isdigit():
            messagebox.showerror("Ошибка", "Введите корректные значения!")
            return
        try:
            quarter = int(quarter)
            parser = KParser()
            result_grade = parser.get_grade((grade, dict(parser.grades).get(grade)), quarter)

            input_file = f'./data/{grade}-{quarter}.xlsx'
            output_file = f'./data/{grade}-{quarter}-result.xlsx'
            calculate_averages(input_file, output_file)

            messagebox.showinfo("Успех", f"Файл сохранен: {output_file}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    tk.Button(root, text="Обработать", command=submit).grid(row=2, column=0, columnspan=2)

    root.mainloop()


def main():
    # Вызов интерфейса
    start_ui()

if __name__ == "__main__":
    main()