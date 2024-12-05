import os
import time
from io import StringIO
from threading import Thread
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pandas import DataFrame

from openpyxl import load_workbook
from pathlib import Path
import re

import tkinter as tk
from tkinter import messagebox, ttk


class RThread(Thread):
    def __init__(self, target, args=()):
        Thread.__init__(self)
        self.target = target
        self.args = args  # Сохраняем аргументы
        self.result = None

    def run(self) -> None:
        self.result = self.target(*self.args)  # Передаем аргументы при вызове целевой функции

class Grade:
    """
    Объект класса.

    Args:
        subjects (dict[str, DataFrame]): урок, таблица
        grade (str): класс
        quarter (int): четверть
    """

    def __init__(self, subjects: dict[str, DataFrame], grade: str, quarter: int):
        self.subjects = subjects
        self.grade = grade
        self.quarter = quarter

    def to_excel(self, path: str = "./data"):
        """
        Создание Excel файла с данными о предмете.
        """
        os.makedirs(path, exist_ok=True)
        path = f"{path}/{self.grade}-{self.quarter}.xlsx"
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            for subject, table in self.subjects.items():
                table.to_excel(writer, sheet_name=subject)

    def print_subjects(self):
        """
        Выввод уроков ввиде таблицы pandas
        """
        for i in self.subjects.items():
            subject = i[0]
            table = i[1]

            print(subject, table, sep="\n", end="\n")


class KParser:
    """
    Класс для работы с системой Kundoluk.

    Основные функции:
    - Получение данных о классах (get_grade)
    - Получение данных по предметам (get_subject)
    - Обработка пользовательского ввода (magic)
    """

    def __init__(self) -> None:
        # Настраиваем заголовки для HTTP-запросов
        self.__headers = {
            "cookie": "language=ru; session=j4jrc9k0qgublh2ih1j6en0vuob9hf2t; deferApp=true; string=6751e3e2be5ae",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "dnt": "1",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "sec-gpc": "1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6324.206 Safari/537.36",
        }

        self.url = "https://kundoluk.edu.kg/journal2"  # Базовый URL для запросов
        self.s = (
            requests.Session()
        )  # Создаем сессию для повторного использования соединений
        # Список классов и их идентификаторов
        self.grades = [
            ("2А", 90359),
            ("2Б", 90360),
            ("3А", 90340),
            ("3Б", 90341),
            ("4А", 90343),
            ("4Б", 90344),
            ("5А", 90345),
            ("5Б", 90346),
            ("6А", 90347),
            ("6Б", 90348),
            ("7А", 90349),
            ("7Б", 90350),
            ("8А", 90351),
            ("8Б", 90352),
            ("9А", 90353),
            ("9Б", 90354),
            ("10А", 90355),
            ("10Б", 90356),
            ("11А", 90357),
            ("11Б", 90358),
            ("2В", 90361),
        ]

    def get_grade(self, grade: tuple[str, int], quarter: int = 1) -> Grade:
        """
        Получение данных для конкретного класса и четверти.

        Returns:
            Grade: Объект Grade с данными класса
        """
        try:
            querystring = {"class": str(grade[1]), "quarter": str(quarter)}
            response = self.s.get(self.url, headers=self.__headers, params=querystring)
            response.raise_for_status()  # Проверка успешности запроса

            bs = BeautifulSoup(response.content, "lxml")
            subjects = bs.find("ul", class_="uk-subnav").find_all("a")
            threads = []
            for link in subjects:
                url = link["href"]
                name = link.text.strip()

                # Создаем поток для получения данных о предмете
                thread = RThread(target=lambda: self.get_subject(url), args=())
                thread.start()
                threads.append((thread, name))

                time.sleep(0.1)  # Небольшая пауза между запросами

            tables = []
            for thread in threads:
                while thread[0].result is None:  # Ожидание завершения потока
                    time.sleep(0.1)
                table = thread[0].result
                if type(table) != int:
                    tables.append((thread[1], table))
            # Сортировка по названию урока
            tables = dict(sorted(tables, key=lambda x: x[0]))

            return Grade(tables, grade[0], quarter)
        except requests.exceptions.RequestException as e:
            raise Exception(f"Ошибка при запросе данных для класса {grade[0]}: {e}")

    def get_subject(self, url: str) -> DataFrame:
        """
        Получение данных о предмете по ссылке.

        Args:
            url (str): Ссылка на предмет

        Returns:
            DataFrame: Таблица с оценками
        """
        response = self.s.request("GET", url, headers=self.__headers)
        # Проверка успешности запроса
        if not response.ok:
            raise Exception(
                "subjectError:",
                f"subject url({url}) response code {response.status_code}",
            )

        bs = BeautifulSoup(response.content, "lxml")

        try:
            htmlTable = bs.find(
                "table", class_="elementFixed-striped"
            )  # Таблица с данными
            # Удаляем лишние элементы из таблицы
            trash = htmlTable.find_all("span", class_="uk-margin-xsmall-right")
            [i.extract() for i in trash]
            table = pd.read_html(StringIO(str(htmlTable)))[
                0
            ]  # Конвертируем в DataFrame
        except Exception:
            return 10
        return table


    def magic(self, grade, quarter) -> None:
        """
        Основная логика обработки пользовательского ввода, полученного через tkinter.
        """
        try:
            # Получаем данные о классе
            grades = dict(self.grades)
            gradeId = grades.get(grade)

            if not gradeId or quarter <= 0:
                print("Класс или четверть указаны неверно.")
                return

            # Получение данных о классе
            grade_data = self.get_grade((grade, gradeId), quarter)

            # Убедитесь, что директория существует
            output_dir = os.path.join(os.getcwd(), 'data')  # Папка 'data' рядом с программой
            os.makedirs(output_dir, exist_ok=True)

            # Создаем путь к файлам
            input_file = os.path.join(output_dir, f'{grade}-{quarter}.xlsx')
            grade_data.to_excel(input_file)

            # Генерация output_file
            output_file = os.path.join(output_dir, f'{grade}-{quarter}-result.xlsx')
            print(f"Генерация файла: {output_file}")

            # Проверка на существование output_file
            if not output_file:
                raise ValueError("Неверный путь для выходного файла.")

            # Обрабатываем данные и сохраняем в новый файл
            calculate_averages(input_file, output_file)

            print(f'Результаты сохранены в файл: {output_file}')

        except Exception as e:
            import traceback
            print(f"Ошибка: {e}")
            print("Traceback:")
            print(traceback.format_exc())  # Вывод полного traceback


def extract_digit(text):
    """Функция для извлечения чисел из текста, поддерживает как целые, так и дробные числа."""
    match = re.search(r'[\d,\.]+', text)
    return float(match.group().replace(',', '.')) if match else 0.0  # Обрабатываем как дробное число


def is_file_open(file_path):
    """Проверяет, открыт ли файл в другой программе."""
    try:
        os.rename(file_path, file_path)  # Пытаемся переименовать файл
        return False  # Если удается, файл не открыт
    except OSError:
        return True  # Если ошибка, значит файл открыт

def check_file_access(file_path):
    """Проверяет, можно ли записывать в файл."""
    try:
        # Пробуем открыть файл на запись
        with open(file_path, 'a'):
            return True
    except PermissionError:
        return False

def calculate_averages(input_file, output_file):
    input_file = Path(input_file).as_posix()

    # Проверка на существование входного файла
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Файл не найден: {input_file}")

    # Проверка, открыт ли файл в другой программе
    if is_file_open(input_file):
        raise PermissionError(f"Файл {input_file} открыт в другой программе. Закройте его и повторите попытку.")

    # Проверка прав доступа для записи в выходной файл
    if not check_file_access(output_file):
        raise PermissionError(f"Нет прав на запись в файл: {output_file}")

    try:
        wb = load_workbook(input_file)

        # Проверка на существование листа 'Result', если уже был создан
        if 'Result' in wb.sheetnames:
            result_sheet = wb['Result']
            wb.remove(result_sheet)

        result_sheet = wb.create_sheet('Result')

        # Объединение ячеек и запись заголовков
        result_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        result_sheet['A1'] = f"Результаты {Path(input_file).stem}"

        subjects = [sheet for sheet in wb.sheetnames if sheet != 'Result']
        for col, subject in enumerate(subjects, start=4):
            result_sheet.cell(row=2, column=col, value=subject)

        result_sheet.cell(row=2, column=len(subjects) + 4, value="Средний балл по всем предметам")

        students = []
        subject_sheet = wb[subjects[0]]
        for i in range(4, subject_sheet.max_row + 1):
            students.append(f"{subject_sheet.cell(row=i, column=1).value} {subject_sheet.cell(row=i, column=2).value}")

        all_averages = {student: [] for student in students}

        for col, subject in enumerate(subjects, start=4):
            subject_sheet = wb[subject]

            sr_col = None
            for col_idx in range(1, subject_sheet.max_column + 1):
                if subject_sheet.cell(row=2, column=col_idx).value == "СР":
                    sr_col = col_idx
                    break

            if sr_col:
                for i, student in enumerate(students, start=4):
                    grade = subject_sheet.cell(row=i, column=sr_col).value
                    grade = extract_digit(str(grade))
                    all_averages[student].append(grade)
                    result_sheet.cell(row=i - 1, column=col, value=grade)

        for row, student in enumerate(students, start=3):
            result_sheet.cell(row=row, column=1, value=row - 2)
            result_sheet.cell(row=row, column=2, value=student)

            total_average = sum(all_averages[student]) / len(subjects)
            result_sheet.cell(row=row, column=len(subjects) + 4, value=round(total_average, 2))

        # Создание нового имени файла с суффиксом '-result'
        input_path = Path(input_file)
        output_file = input_path.with_name(input_path.stem + '-result' + input_path.suffix)

        # Преобразование пути для сохранения в формат с прямыми слэшами
        output_file = output_file.as_posix()

        # Сохранение файла
        wb.save(output_file)
        print(f"Файл сохранен: {output_file}")

    except PermissionError as e:
        print(f"Ошибка доступа: {e}")
    except FileNotFoundError as e:
        print(f"Файл не найден: {e}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


def start_ui():
    # Создание главного окна
    root = tk.Tk()
    root.title("Kundoluk Parser")

    # Настройка стилей для ttk
    style = ttk.Style()
    style.configure("RoundedButton.TButton",
                    background="#3C3E52",  # Цвет фона по умолчанию
                    foreground="#FFFFFF",  # Цвет текста по умолчанию
                    font=("Arial", 14),  # Кастомный шрифт
                    padding=10,
                    bd=2)

    # Настройки для изменения кнопки при наведении
    style.map("RoundedButton.TButton",
              background=[("active", "#FFFFFF"),  # Цвет фона при наведении (светлый)
                          ("!active", "#333333")],  # Цвет фона в неактивном состоянии (темный)
              foreground=[("active", "#FFFFFF"),  # Цвет текста при наведении (темный)
                          ("!active", "#333333")]),  # Цвет текста в неактивном состоянии (белый)

    # Установка курсора для кнопки
    style.map("RoundedButton.TButton",
              cursor=[("active", "hand2")])  # Изменяет курсор на "hand2" (рука при наведении)

    # Установка фоновых цветов для окна
    root.config(bg="#F6F6F6")

    # Заголовки
    tk.Label(root, text="Введите класс (например, 4Б):", font=("Arial", 14, "bold"), bg="#F6F6F6", fg="#3C3E52").grid(row=0,
                                                                                                              column=0,
                                                                                                              padx=10,
                                                                                                              pady=10)
    tk.Label(root, text="Введите четверть (число):", font=("Arial", 14, "bold"), bg="#F6F6F6", fg="#3C3E52").grid(row=1,
                                                                                                          column=0,
                                                                                                          padx=10,
                                                                                                          pady=10)

    # Поля ввода
    grade_entry = tk.Entry(root, font=("Arial", 12), bd=2, relief="solid", width=20)
    quarter_entry = tk.Entry(root, font=("Arial", 12), bd=2, relief="solid", width=20)

    grade_entry.grid(row=0, column=1, padx=10, pady=10)
    quarter_entry.grid(row=1, column=1, padx=10, pady=10)

    # Функция для обработки отправки
    def submit():
        grade = grade_entry.get().strip().upper()
        quarter = quarter_entry.get().strip()

        if not grade or not quarter.isdigit():
            messagebox.showerror("Ошибка", "Введите корректные значения!")
            return

        try:
            quarter = int(quarter)
            parser = KParser()
            parser.magic(grade, quarter)  # передаем данные в magic()

            # Генерация выходного файла
            input_file = f'./data/{grade}-{quarter}.xlsx'
            output_file = f'./data/{grade}-{quarter}-result.xlsx'
            calculate_averages(input_file, output_file)

            messagebox.showinfo("Успех", f"Файл сохранен: {output_file}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    # Кнопка с округленными углами и изменением стилей при наведении
    submit_button = ttk.Button(root, text="Отправить", style="RoundedButton.TButton", command=submit)
    submit_button.grid(row=2, column=0, columnspan=2, pady=5)

    root.mainloop()

def main():
    # parser = KParser()
    # parser.magic()  # Запуск основной логики
    start_ui()

if __name__ == "__main__":
    main()


