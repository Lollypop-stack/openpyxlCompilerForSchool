import os
import time
import openpyxl
from io import StringIO
from threading import Thread
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pandas import DataFrame

from openpyxl.chart import PieChart, Reference

from openpyxl import load_workbook
from pathlib import Path
import re

import tkinter as tk
from tkinter import messagebox, ttk



class RThread(Thread):
    def __init__(self, target, args=()):
        Thread.__init__(self)
        self.target = target
        self.args = args  # Сохраняем аргументы
        self.result = None

    def run(self) -> None:
        self.result = self.target(*self.args)  # Передаем аргументы при вызове целевой функции

class Grade:
    """
    Класс для работы с данными класса, содержащий предметы и четверть.
    """
    def __init__(self, subjects: dict[str, pd.DataFrame], grade: str, quarter: int):
        self.subjects = subjects
        self.grade = grade
        self.quarter = quarter

    def to_excel(self, path: str):
        """Создание Excel файла с данными по предметам в указанном пути."""
        folder_name = f"{self.grade}-{self.quarter}"
        folder_path = os.path.join(path, f"data/{self.grade}-{self.quarter}")
        os.makedirs(folder_path, exist_ok=True)

        file_path = os.path.join(path, f"data/{self.grade}-{self.quarter}/{self.grade}-{self.quarter}.xlsx")
        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            for subject, table in self.subjects.items():
                table.to_excel(writer, sheet_name=subject)
        return file_path

    def print_subjects(self):
        """
        Вывод предметов в консоль.
        """
        for subject, table in self.subjects.items():
            print(subject, table, sep="\n", end="\n")

class KParser:
    """
    Класс для работы с системой Kundoluk.

    Основные функции:
    - Получение данных о классах (get_grade)
    - Получение данных по предметам (get_subject)
    - Обработка пользовательского ввода (magic)
    """

    def __init__(self) -> None:
        # Настраиваем заголовки для HTTP-запросов
        self.__headers = {
            "cookie": "language=ru; session=f2hnvqugvmgo3dnp997dip0fe90mjblb; deferApp=true; string=6751e3e2be5ae",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "dnt": "1",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "sec-gpc": "1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6324.206 Safari/537.36",
            "response" : "self.s.get(self.url, headers=self.__headers, params=querystring, timeout=10)",
        }

        self.url = "https://kundoluk.edu.kg/journal2"  # Базовый URL для запросов
        self.s = (
            requests.Session()
        )  # Создаем сессию для повторного использования соединений
        # Список классов и их идентификаторов
        self.grades = [
            ("2А", 90359),
            ("2Б", 90360),
            ("3А", 90340),
            ("3Б", 90341),
            ("4А", 90343),
            ("4Б", 90344),
            ("5А", 90345),
            ("5Б", 90346),
            ("6А", 90347),
            ("6Б", 90348),
            ("7А", 90349),
            ("7Б", 90350),
            ("8А", 90351),
            ("8Б", 90352),
            ("9А", 90353),
            ("9Б", 90354),
            ("10А", 90355),
            ("10Б", 90356),
            ("11А", 90357),
            ("11Б", 90358),
            ("2В", 90361),
        ]

    def get_grade(self, grade: tuple[str, int], quarter: int = 1) -> Grade:
        """
        Получение данных для конкретного класса и четверти.
        Returns:
            Grade: Объект Grade с данными класса
        """
        try:
            querystring = {"class": str(grade[1]), "quarter": str(quarter)}
            response = self.s.get(self.url, headers=self.__headers, params=querystring)
            response.raise_for_status()

            bs = BeautifulSoup(response.content, "lxml")
            subjects = bs.find("ul", class_="uk-subnav").find_all("a")
            threads = []
            for link in subjects:
                url = link["href"]
                name = link.text.strip()
                thread = RThread(target=lambda: self.get_subject(url), args=())
                thread.start()
                threads.append((thread, name))

                time.sleep(0.1)

            tables = []
            for thread in threads:
                thread[0].join()  # Дождаться завершения потока
                table = thread[0].result
                if type(table) != int:
                    tables.append((thread[1], table))
            tables = dict(sorted(tables, key=lambda x: x[0]))

            return Grade(tables, grade[0], quarter)
        except requests.exceptions.RequestException as e:
            raise Exception(f"Ошибка при запросе данных для класса {grade[0]}: {e}")

    def get_subject(self, url: str) -> pd.DataFrame:
        """Получение данных о предмете по ссылке."""
        response = self.s.request("GET", url, headers=self.__headers)
        if not response.ok:
            raise Exception(f"Ошибка запроса для {url} с кодом {response.status_code}")

        bs = BeautifulSoup(response.content, "lxml")
        try:
            htmlTable = bs.find("table", class_="elementFixed-striped")
            trash = htmlTable.find_all("span", class_="uk-margin-xsmall-right")
            [i.extract() for i in trash]
            table = pd.read_html(StringIO(str(htmlTable)))[0]
        except Exception:
            return pd.DataFrame()  # Пустой DataFrame при ошибке
        return table

    def magic(self, grade, quarter) -> None:
        """Основная логика обработки пользовательского ввода."""
        try:
            grades = dict(self.grades)
            gradeId = grades.get(grade)

            if not gradeId or quarter <= 0:
                print("Класс или четверть указаны неверно.")
                return

            grade_data = self.get_grade((grade, gradeId), quarter)

            # Преобразование данных: заменяем нули на пустые клетки
            for subject, table in grade_data.subjects.items():
                grade_data.subjects[subject] = table.replace(0, '')

            output_dir = os.getcwd()  # Основная папка для сохранения
            input_file = grade_data.to_excel(output_dir)  # Сохранение данных в файл с преобразованием

            print(f"Данные сохранены в файл: {input_file}")

            # Путь к файлу, который будем изменять
            output_file = input_file
            print(f"Генерация результата в файл: {output_file}")

            if not output_file:
                raise ValueError("Неверный путь для выходного файла.")

            calculate_averages(input_file, output_file)  # Расчет средних баллов и сохранение в тот же файл

            print(f'Результаты сохранены в файл: {output_file}')

        except Exception as e:
            import traceback
            print(f"Ошибка: {e}")
            print("Traceback:")
            print(traceback.format_exc())


def extract_digit(text):
    """Функция для извлечения чисел из текста, поддерживает как целые, так и дробные числа."""
    match = re.search(r'[\d,\.]+', text)
    return float(match.group().replace(',', '.')) if match else 0.0  # Обрабатываем как дробное число

def assign_categories(result_sheet, num_subjects):
    """
    Добавляет категории (Отл., Уд., Тр., Дв., Нз.) рядом со средним баллом.

    Аргументы:
    result_sheet -- лист Excel с результатами
    num_subjects -- количество предметов
    """
    category_column = num_subjects + 5  # Колонка для записи категорий
    result_sheet.cell(row=2, column=category_column, value="Категория")

    for row in range(3, result_sheet.max_row + 1):
        avg_cell = result_sheet.cell(row=row, column=num_subjects + 4)
        avg_score = avg_cell.value

        if avg_score is None:
            category = "Нз."  # Незачет, если средний балл отсутствует
        elif avg_score > 4.6:
            category = "Отл."  # Отличник
        elif avg_score > 3.6:
            category = "Уд."  # Ударник
        elif avg_score > 2.6:
            category = "Тр."  # Троечник
        else:
            category = "Дв."  # Двоечник

        result_sheet.cell(row=row, column=category_column, value=category)


def create_category_pie_chart(result_sheet, category_column, start_row, end_row, num_subjects):
    """
    Создает круговую диаграмму по категориям студентов (Отл., Уд., Тр., Дв., Нз.)
    Мини-таблица с категориями будет отображаться справа от столбца с предметами.

    Аргументы:
    result_sheet -- лист Excel с результатами
    category_column -- номер колонки с категориями
    start_row -- начальная строка для анализа категорий
    end_row -- последняя строка для анализа категорий
    num_subjects -- количество предметов
    """
    # Подсчитываем количество студентов в каждой категории
    categories = ["Отл.", "Уд.", "Тр.", "Дв.", "Нз."]
    category_counts = {category: 0 for category in categories}

    # Собираем данные о категориях
    for row in range(start_row, end_row + 1):
        category_cell = result_sheet.cell(row=row, column=category_column)
        category = category_cell.value
        if category in category_counts:
            category_counts[category] += 1

    # Записываем мини-таблицу с категориями
    table_start_col = num_subjects + 7
    table_start_row = 3  # Мини-таблица начинается с 3 строки
    result_sheet.cell(row=2, column=table_start_col, value="Категория")
    result_sheet.cell(row=2, column=table_start_col + 1, value="Кол-во")
    result_sheet.cell(row=2, column=table_start_col + 2, value="Процент")

    total_students = sum(category_counts.values())
    for i, category in enumerate(categories, start=table_start_row):
        result_sheet.cell(row=i, column=table_start_col, value=category)
        result_sheet.cell(row=i, column=table_start_col + 1, value=category_counts[category])
        percent = (category_counts[category] / total_students * 100) if total_students > 0 else 0
        result_sheet.cell(row=i, column=table_start_col + 2, value=f"{percent:.2f}%")

    # Создаем круговую диаграмму
    chart = PieChart()
    data = Reference(result_sheet, min_col=table_start_col + 1, min_row=table_start_row,
                     max_row=table_start_row + len(categories) - 1)
    categories_reference = Reference(result_sheet, min_col=table_start_col, min_row=table_start_row,
                                     max_row=table_start_row + len(categories) - 1)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories_reference)
    chart.title = "Распределение категорий"

    # Позиция диаграммы (под мини-таблицей)
    chart_anchor = f"{openpyxl.utils.get_column_letter(table_start_col)}{table_start_row + len(categories) + 1}"
    result_sheet.add_chart(chart, chart_anchor)

def calculate_averages(input_file, output_file):
    try:
        wb = load_workbook(input_file)

        # Удаляем старый лист 'Result', если он существует
        if 'Result' in wb.sheetnames:
            wb.remove(wb['Result'])

        # Создаём новый лист 'Result'
        result_sheet = wb.create_sheet('Result')

        # Объединяем ячейки для заголовка и записываем его
        result_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        result_sheet['A1'] = f"Результаты {Path(input_file).stem}"

        # Собираем список предметов (названий листов, кроме 'Result')
        subjects = [sheet for sheet in wb.sheetnames if sheet != 'Result']
        num_subjects = len(subjects)  # Подсчитаем количество предметов

        # Записываем заголовки в 'Result'
        for col, subject in enumerate(subjects, start=4):
            result_sheet.cell(row=2, column=col, value=subject)

        result_sheet.cell(row=2, column=len(subjects) + 4, value="Средний балл по всем предметам")


        # Собираем список студентов из первого предмета
        students = []
        first_subject_sheet = wb[subjects[0]]
        for i in range(4, first_subject_sheet.max_row + 1):
            last_name = first_subject_sheet.cell(row=i, column=1).value
            first_name = first_subject_sheet.cell(row=i, column=2).value
            students.append(f"{last_name} {first_name}")

        # Инициализируем словарь для хранения оценок по каждому студенту
        all_averages = {student: [] for student in students}

        # Обрабатываем каждый предмет
        for col, subject in enumerate(subjects, start=4):
            subject_sheet = wb[subject]

            # Ищем колонку 'СР' (средний балл)
            sr_col = next((col_idx for col_idx in range(1, subject_sheet.max_column + 1)
                           if subject_sheet.cell(row=2, column=col_idx).value == "СР"), None)

            if sr_col:
                for i, student in enumerate(students, start=4):
                    grade = subject_sheet.cell(row=i, column=sr_col).value
                    if grade:
                        grade = float(str(grade).replace(",", "."))
                        all_averages[student].append(grade)
                        result_sheet.cell(row=i - 1, column=col, value=grade)

        # Заполняем итоговые данные по студентам
        for row, student in enumerate(students, start=3):
            result_sheet.cell(row=row, column=1, value=row - 2)
            result_sheet.cell(row=row, column=2, value=student)

            # Считаем средний балл по всем предметам без учёта нулей
            grades = all_averages[student]
            total_average = sum(grades) / len(grades) if grades else 0
            result_sheet.cell(row=row, column=len(subjects) + 4, value=round(total_average, 2))

        assign_categories(result_sheet, len(subjects))

        create_category_pie_chart(result_sheet, num_subjects + 5, 3, len(students) + 2, len(subjects))

        # Сохраняем обновлённый файл
        wb.save(output_file)

    except PermissionError as e:
        print(f"Ошибка доступа: {e}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")

def is_file_open(file_path):
    """Проверяет, открыт ли файл в другой программе."""
    try:
        os.rename(file_path, file_path)  # Пытаемся переименовать файл
        return False  # Если удается, файл не открыт
    except OSError:
        return True  # Если ошибка, значит файл открыт

def check_file_access(file_path):
    """Проверяет, можно ли записывать в файл."""
    try:
        # Пробуем открыть файл на запись
        with open(file_path, 'a'):
            return True
    except PermissionError:
        return False



def perform_task(grade, quarter):
    try:
        # Создаем папку и файл для класса
        data_dir = os.path.join(os.getcwd(), f'data/{grade}-{quarter}')
        os.makedirs(data_dir, exist_ok=True)

        # Загружаем данные для класса
        parser = KParser()  # В вашем примере, это код для парсинга данных
        grade_data = parser.get_grade((grade, quarter), quarter)

        input_file = grade_data.to_excel(data_dir)  # Сохранение данных в файл

        # Создаем файл для результатов
        output_file = input_file.replace(".xlsx", "-result.xlsx")
        calculate_averages(input_file, output_file)  # Расчет и сохранение результатов

    except Exception as e:
        print(f"Ошибка: {e}")

    except Exception as e:
        print(f"Ошибка: {e}")


def center_window(window, width, height):
    """Центрирование окна на экране"""
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)

    window.geometry(f'{width}x{height}+{position_right}+{position_top}')


def start_ui():
    # Создание главного окна
    root = tk.Tk()
    root.title("Kundoluk Parser")

    # Настройка стилей для ttk
    style = ttk.Style()
    style.configure("RoundedButton.TButton",
                    background="#3C3E52",  # Цвет фона по умолчанию
                    foreground="#FFFFFF",  # Цвет текста по умолчанию
                    font=("Arial", 14),  # Кастомный шрифт
                    padding=10,
                    bd=2)

    center_window(root, 550, 160)  # Ширина 400, высота 25

    # Настройки для изменения кнопки при наведении
    style.map("RoundedButton.TButton",
              background=[("active", "#FFFFFF"),  # Цвет фона при наведении (светлый)
                          ("!active", "#333333")],  # Цвет фона в неактивном состоянии (темный)
              foreground=[("active", "#FFFFFF"),  # Цвет текста при наведении (темный)
                          ("!active", "#333333")]),  # Цвет текста в неактивном состоянии (белый)

    # Установка курсора для кнопки
    style.map("RoundedButton.TButton",
              cursor=[("active", "hand2")])  # Изменяет курсор на "hand2" (рука при наведении)

    # Установка фоновых цветов для окна
    root.config(bg="#F6F6F6")

    # Заголовки
    tk.Label(root, text="Введите класс (например, 4Б):", font=("Arial", 14, "bold"), bg="#F6F6F6", fg="#3C3E52").grid(
        row=0,
        column=0,
        padx=10,
        pady=10)
    tk.Label(root, text="Введите четверть (число):", font=("Arial", 14, "bold"), bg="#F6F6F6", fg="#3C3E52").grid(row=1,
                                                                                                                  column=0,
                                                                                                                  padx=10,
                                                                                                                  pady=10)

    # Поля ввода
    grade_entry = tk.Entry(root, font=("Arial", 12), bd=2, relief="solid", width=20)
    quarter_entry = tk.Entry(root, font=("Arial", 12), bd=2, relief="solid", width=20)

    grade_entry.grid(row=0, column=1, padx=10, pady=10)
    quarter_entry.grid(row=1, column=1, padx=10, pady=10)

    def show_preloader():
        # Здесь создаем окно прелоадера с зеленой полосой загрузки
        preloader = tk.Toplevel()
        preloader.title("Загрузка...")
        center_window(preloader, 300, 120)  # Ширина 300, высота

        # Элемент текста
        label = tk.Label(preloader, text="Пожалуйста, подождите...", font=("Arial", 12), bg="#F6F6F6", fg="#3C3E52")
        label.pack(padx=20, pady=10)

        # Зеленая полоска загрузки
        progress = ttk.Progressbar(preloader, length=250, mode='indeterminate', style="TProgressbar")
        progress.pack(padx=20, pady=10)

        # Начинаем анимацию полосы
        progress.start()

        return preloader, progress

    def perform_task(grade, quarter, preloader):
        try:
            # Ваш код для обработки данных
            parser = KParser()
            parser.magic(grade, quarter)

            input_file = f'./data/{grade}-{quarter}.xlsx'
            output_file = f'./data/{grade}-{quarter}-result.xlsx'
            calculate_averages(input_file, output_file)

        finally:
            preloader.destroy()  # Закрываем прелоадер после завершения

    def submit_with_preloader():
        """
        Обработчик кнопки "Отправить" с прелоадером.
        """
        grade = grade_entry.get().strip().upper()
        quarter = quarter_entry.get().strip()

        if not grade or not quarter.isdigit():
            messagebox.showerror("Ошибка", "Введите корректные значения!")
            return

        quarter = int(quarter)

        # Создаем окно прелоадера
        preloader, progress = show_preloader()

        # Выполняем задачу в отдельном потоке
        task_thread = Thread(target=perform_task, args=(grade, quarter, preloader))
        task_thread.start()

    # Кнопка с округленными углами и изменением стилей при наведении
    submit_button = ttk.Button(root, text="Отправить", style="RoundedButton.TButton", command=submit_with_preloader)
    submit_button.grid(row=2, column=0, columnspan=2, pady=5)

    root.mainloop()

def main():
    start_ui()

if __name__ == "__main__":
    main()



