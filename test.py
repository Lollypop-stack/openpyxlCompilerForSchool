import re
from pathlib import Path
from openpyxl import load_workbook

def extract_digit(text):
    """Функция для извлечения чисел из текста, поддерживает как целые, так и дробные числа."""
    match = re.search(r'[\d,\.]+', text)
    return float(match.group().replace(',', '.')) if match else 0.0  # Обрабатываем как дробное число

def calculate_averages(input_file, output_file):
    wb = load_workbook(input_file)

    # Создаем лист для результатов
    result_sheet = wb.create_sheet('Result')


    # Записываем название класса в объединенную ячейку A1
    subjects = [subject for subject in wb.sheetnames if subject != 'Result']
    result_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subjects) + 3)

    # ////////////////////////////////////////////////
    input_path = Path(input_file)
    file_name = input_path.stem
    result_sheet['A1'] = f"{file_name} результаты"
    # ////////////////////////////////////////////////

    # Устанавливаем заголовки "Ученик" (объединенные B2 и C2)
    result_sheet.merge_cells('B2:C2')
    result_sheet['B2'] = 'Ученик'

    # Записываем названия предметов в строку 2
    for col, subject in enumerate(subjects, start=4):
        result_sheet.cell(row=2, column=col, value=subject)

    # Заголовок для средней оценки по всем предметам
    result_sheet.cell(row=2, column=len(subjects) + 4, value="Средний балл по всем предметам")

    # Получение списка учеников с первого листа
    subject_sheet = wb[subjects[0]]
    students = [
        f"{str(subject_sheet.cell(row=i, column=1).value)} {str(subject_sheet.cell(row=i, column=2).value)}"
        for i in range(4, subject_sheet.max_row + 1)
    ]

    # Массив для хранения всех оценок по всем предметам для каждого ученика
    all_averages = {student: [] for student in students}

    # Пройдемся по каждому листу, чтобы получить оценки
    for col, subject_name in enumerate(subjects, start=4):
        subject_sheet = wb[subject_name]

        # Поиск колонки с текстом "СР" (предположим, что она в 2-й строке)
        sr_col = None
        for col_idx in range(1, subject_sheet.max_column + 1):
            if subject_sheet.cell(row=2, column=col_idx).value == "СР":  # Ищем колонку "СР"
                sr_col = col_idx
                break

        # Если нашли колонку "СР", копируем оценки для всех учеников
        if sr_col:
            # Записываем название предмета в строку 3 в колонку результата
            result_sheet.cell(row=3, column=col, value=subject_name)

            # Копируем средние баллы для каждого ученика
            for i, student in enumerate(students, start=4):
                grade = subject_sheet.cell(row=i, column=sr_col).value
                grade = extract_digit(str(grade))  # Извлекаем число из строки, теперь поддерживаем дробные числа
                all_averages[student].append(grade)  # Добавляем оценку в общий список
                result_sheet.cell(row=i - 1, column=col, value=grade)  # Записываем оценку в итоговую таблицу (сдвиг на 1 вниз)

    # Заполняем средние оценки по всем предметам для каждого ученика
    for row, student in enumerate(students, start=3):
        result_sheet.cell(row=row, column=1, value=row - 2)  # Номер ученика
        result_sheet.cell(row=row, column=2, value=student)

        # Рассчитываем средний балл по всем предметам для каждого ученика (с округлением)
        total_average = sum(all_averages[student]) / len(subjects)
        result_sheet.cell(row=row, column=len(subjects) + 4, value=round(total_average, 2))  # Округление для среднего балла по всем предметам

    # Сохраняем новый файл
    wb.save(output_file)

# Пример использования функции
input_file = '8А-1.xlsx'  # Исходный файл с оценками
output_file = 'fin.xlsx'  # Новый файл с результатами
calculate_averages(input_file, output_file)
